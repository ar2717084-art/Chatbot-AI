from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from dotenv import load_dotenv
import os
import requests
import uuid
import time
import logging
import json
import base64
from typing import Optional, List
from pathlib import Path
import io

load_dotenv()

# ─── Logging ───────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("nova")

# ─── App ───────────────────────────────────────────────────
app = FastAPI(title="Nova AI Backend", version="3.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"
MODEL = "llama-3.3-70b-versatile"   # upgraded to 70B for better quality
MAX_HISTORY = 40                     # messages kept per session
SESSION_TTL = 7200                   # 2 hours idle timeout

# ─── Session store ─────────────────────────────────────────
# { session_id: { "history": [...], "last_used": float, "title": str, "created": float } }
sessions: dict[str, dict] = {}

# ─── Persistent chat history (file-based) ──────────────────
HISTORY_DIR = Path("chat_histories")
HISTORY_DIR.mkdir(exist_ok=True)

def save_session_to_disk(session_id: str, session: dict):
    """Persist session to disk for history feature."""
    try:
        path = HISTORY_DIR / f"{session_id}.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(session, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log.warning(f"Could not save session: {e}")

def load_session_from_disk(session_id: str) -> Optional[dict]:
    """Load session from disk."""
    try:
        path = HISTORY_DIR / f"{session_id}.json"
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        log.warning(f"Could not load session: {e}")
    return None

def list_all_sessions() -> list:
    """Return all saved sessions metadata for sidebar display."""
    results = []
    for path in sorted(HISTORY_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            results.append({
                "session_id": path.stem,
                "title": data.get("title", "New Conversation"),
                "created": data.get("created", 0),
                "last_used": data.get("last_used", 0),
                "message_count": len(data.get("history", [])),
            })
        except:
            pass
    return results[:50]  # return most recent 50

# ─── System prompt ─────────────────────────────────────────
SYSTEM_PROMPT = """You are Nova, a highly capable, smart, and friendly AI assistant — similar to Claude or ChatGPT in quality.

Core Principles:
- Be accurate, thoughtful, and genuinely helpful
- Give complete, well-structured answers unless brevity is clearly appropriate
- Think step-by-step for complex problems before answering
- If you don't know something, say so honestly — never fabricate facts
- Respond in the same language the user writes in

Formatting:
- Use markdown naturally: headers for long content, **bold** for key terms, bullet points when listing items
- Use code blocks with proper language tags for all code
- For step-by-step tasks, use numbered lists
- Keep paragraphs concise — 3-4 sentences max each

Tone:
- Warm, professional, and direct
- Conversational for casual questions, thorough for technical ones
- Never overly verbose or padding responses with filler

When files are provided:
- Read and analyze them carefully
- Extract key information, structure, and insights
- Reference specific parts of the file content in your response
- If it's code, analyze logic and suggest improvements
- If it's a document, summarize and answer questions about it
"""

# ─── Mode instructions ─────────────────────────────────────
MODE_INSTRUCTIONS = {
    "chat":      lambda msg: msg,
    "info":      lambda msg: f"Explain clearly and thoroughly, using examples and analogies where helpful. Structure your explanation with clear sections:\n\n{msg}",
    "translate": lambda msg: f"Detect the source language and translate to English (or if English, translate to Urdu). Format:\n**Detected Language:** [language]\n**Translation:**\n[translation]\n\nText to translate:\n{msg}",
    "suggest":   lambda msg: f"Give 5 specific, practical, and actionable suggestions for the following. For each: provide a title, brief explanation, and one concrete next step:\n\n{msg}",
    "game":      lambda msg: f"You are an engaging game master. Run an interactive text-based game. Keep it fun, descriptive, and give the user clear choices. Current input: {msg}",
    "create":    lambda msg: f"Create high-quality, original content for this request. Be creative, polished, and avoid clichés:\n\n{msg}",
    "summarize": lambda msg: f"Provide a comprehensive summary with:\n- **Key Points** (3-5 bullets)\n- **Main Takeaway** (1-2 sentences)\n- **Notable Details** (anything important not in key points)\n\nText:\n{msg}",
    "code":      lambda msg: f"You are an expert software engineer. Provide clean, well-commented, production-quality code with explanations. Include error handling and best practices:\n\n{msg}",
}

# ─── File reading utilities ─────────────────────────────────
def extract_text_from_file(filename: str, content: bytes) -> str:
    """Extract readable text from various file types."""
    ext = filename.lower().split('.')[-1]
    
    try:
        if ext == 'txt':
            return content.decode('utf-8', errors='replace')
        
        elif ext == 'csv':
            text = content.decode('utf-8', errors='replace')
            lines = text.split('\n')
            preview = '\n'.join(lines[:100])  # first 100 rows
            if len(lines) > 100:
                preview += f"\n... ({len(lines) - 100} more rows)"
            return f"CSV File ({len(lines)} rows):\n{preview}"
        
        elif ext == 'json':
            text = content.decode('utf-8', errors='replace')
            try:
                data = json.loads(text)
                return f"JSON File:\n{json.dumps(data, indent=2, ensure_ascii=False)[:8000]}"
            except:
                return text[:8000]
        
        elif ext == 'md':
            return content.decode('utf-8', errors='replace')
        
        elif ext == 'py':
            return f"Python file:\n```python\n{content.decode('utf-8', errors='replace')}\n```"
        
        elif ext in ('js', 'ts', 'jsx', 'tsx'):
            return f"JavaScript/TypeScript file:\n```javascript\n{content.decode('utf-8', errors='replace')}\n```"
        
        elif ext in ('html', 'htm'):
            return f"HTML file:\n```html\n{content.decode('utf-8', errors='replace')[:8000]}\n```"
        
        elif ext == 'css':
            return f"CSS file:\n```css\n{content.decode('utf-8', errors='replace')[:8000]}\n```"
        
        elif ext == 'pdf':
            # Try extracting with pypdf2 or pdfminer
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    text_parts = []
                    for i, page in enumerate(pdf.pages[:20]):  # max 20 pages
                        text = page.extract_text()
                        if text:
                            text_parts.append(f"--- Page {i+1} ---\n{text}")
                    return f"PDF Document ({len(pdf.pages)} pages):\n\n" + '\n\n'.join(text_parts)
            except ImportError:
                try:
                    import PyPDF2
                    reader = PyPDF2.PdfReader(io.BytesIO(content))
                    text_parts = []
                    for i, page in enumerate(reader.pages[:20]):
                        text = page.extract_text()
                        if text:
                            text_parts.append(f"--- Page {i+1} ---\n{text}")
                    return f"PDF Document ({len(reader.pages)} pages):\n\n" + '\n\n'.join(text_parts)
                except ImportError:
                    return "PDF file uploaded. Note: Install pdfplumber or PyPDF2 for full text extraction."
        
        elif ext in ('doc', 'docx'):
            try:
                import docx
                doc = docx.Document(io.BytesIO(content))
                text = '\n'.join([para.text for para in doc.paragraphs if para.text.strip()])
                return f"Word Document:\n{text[:8000]}"
            except ImportError:
                return "Word document uploaded. Install python-docx for text extraction."
        
        elif ext in ('xlsx', 'xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                result = []
                for sheet_name in wb.sheetnames[:3]:  # max 3 sheets
                    ws = wb[sheet_name]
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= 100: break
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        rows.append('\t'.join(row_data))
                    result.append(f"Sheet: {sheet_name}\n" + '\n'.join(rows))
                return f"Excel File:\n\n" + '\n\n'.join(result)
            except ImportError:
                return "Excel file uploaded. Install openpyxl for data extraction."
        
        elif ext in ('png', 'jpg', 'jpeg', 'gif', 'webp'):
            size_kb = len(content) / 1024
            return f"[Image file: {filename}, {size_kb:.1f} KB — image analysis requires vision model]"
        
        else:
            # Try as plain text
            try:
                return content.decode('utf-8', errors='replace')[:5000]
            except:
                return f"[Binary file: {filename}, {len(content)} bytes]"
    
    except Exception as e:
        log.error(f"File extraction error for {filename}: {e}")
        return f"[Could not read file: {filename} — {str(e)}]"

# ─── Session helpers ────────────────────────────────────────
def clean_sessions():
    now = time.time()
    expired = [sid for sid, data in sessions.items()
               if now - data["last_used"] > SESSION_TTL]
    for sid in expired:
        del sessions[sid]
    if expired:
        log.info(f"Cleaned {len(expired)} expired session(s)")

def get_session(session_id: str) -> dict:
    clean_sessions()
    if session_id not in sessions:
        # Try loading from disk first
        saved = load_session_from_disk(session_id)
        if saved:
            sessions[session_id] = saved
            log.info(f"Restored session from disk: {session_id[:8]}…")
        else:
            sessions[session_id] = {
                "history": [],
                "last_used": time.time(),
                "created": time.time(),
                "title": "New Conversation"
            }
            log.info(f"New session: {session_id[:8]}…")
    sessions[session_id]["last_used"] = time.time()
    return sessions[session_id]

def generate_title(first_message: str) -> str:
    """Generate a short title from the first user message."""
    words = first_message.strip().split()
    title = ' '.join(words[:6])
    if len(words) > 6:
        title += '…'
    return title[:50] or "New Conversation"

def call_groq(messages: list, temperature: float = 0.7) -> str:
    if not GROQ_API_KEY:
        raise EnvironmentError("GROQ_API_KEY is not set in .env")

    response = requests.post(
        GROQ_URL,
        headers={
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "model": MODEL,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": 2048,  # increased from 1024
        },
        timeout=45  # increased timeout
    )

    if response.status_code != 200:
        log.error(f"Groq error {response.status_code}: {response.text[:300]}")
        raise RuntimeError(f"Groq API error: {response.status_code}")

    data = response.json()
    if "choices" not in data or not data["choices"]:
        raise RuntimeError("Unexpected Groq response format")

    return data["choices"][0]["message"]["content"].strip()

# ─── Routes ────────────────────────────────────────────────
@app.get("/")
def home():
    return {
        "status": "Nova AI is running 🚀",
        "version": "3.0",
        "model": MODEL,
        "active_sessions": len(sessions)
    }

@app.get("/health")
def health():
    return {"ok": True, "sessions": len(sessions)}

@app.get("/history")
def get_history():
    """Return list of all saved chat sessions for sidebar."""
    return {"sessions": list_all_sessions()}

@app.get("/session/{session_id}")
def get_session_messages(session_id: str):
    """Return full message history for a session."""
    session = get_session(session_id)
    return {
        "session_id": session_id,
        "title": session.get("title", "New Conversation"),
        "history": session.get("history", []),
        "created": session.get("created", 0),
        "message_count": len(session.get("history", []))
    }

@app.delete("/session/{session_id}")
def clear_session(session_id: str):
    if session_id in sessions:
        del sessions[session_id]
    # Also remove from disk
    path = HISTORY_DIR / f"{session_id}.json"
    if path.exists():
        path.unlink()
    return {"cleared": True}

@app.delete("/history")
def clear_all_history():
    """Clear all saved sessions."""
    sessions.clear()
    for path in HISTORY_DIR.glob("*.json"):
        path.unlink()
    return {"cleared": True}

# ─── Chat with file upload support ─────────────────────────
@app.post("/chat")
async def chat(
    message: str = Form(...),
    mode: str = Form("chat"),
    session_id: str = Form(None),
    files: List[UploadFile] = File(default=[])
):
    start = time.time()

    try:
        # Resolve session
        sid = session_id or str(uuid.uuid4())
        session = get_session(sid)
        history: list = session["history"]

        # Validate mode
        mode = mode if mode in MODE_INSTRUCTIONS else "chat"

        # Process uploaded files
        file_contents = []
        for upload in files:
            if upload.filename:
                raw = await upload.read()
                extracted = extract_text_from_file(upload.filename, raw)
                file_contents.append({
                    "name": upload.filename,
                    "content": extracted
                })
                log.info(f"Processed file: {upload.filename} ({len(raw)} bytes → {len(extracted)} chars)")

        # Build user message
        build = MODE_INSTRUCTIONS[mode]
        user_text = message.strip()
        
        # Combine message + file contents
        full_user_input = build(user_text)
        if file_contents:
            file_section = "\n\n".join([
                f"=== File: {f['name']} ===\n{f['content']}"
                for f in file_contents
            ])
            full_user_input = f"{full_user_input}\n\n--- Attached Files ---\n{file_section}"

        # Set session title from first real message
        if not history and user_text:
            session["title"] = generate_title(user_text)

        # Append to history
        history.append({"role": "user", "content": full_user_input})

        # Trim history
        if len(history) > MAX_HISTORY:
            history = history[-MAX_HISTORY:]
        session["history"] = history

        # Groq messages
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            *history
        ]

        # Adjust temperature by mode
        temp_map = {
            "chat": 0.7, "info": 0.4, "translate": 0.2,
            "suggest": 0.8, "game": 0.9, "create": 0.85,
            "summarize": 0.3, "code": 0.2
        }

        reply = call_groq(messages, temperature=temp_map.get(mode, 0.7))

        # Save reply to history
        history.append({"role": "assistant", "content": reply})
        session["history"] = history

        # Persist to disk
        save_session_to_disk(sid, session)

        elapsed = round((time.time() - start) * 1000)
        log.info(f"[{mode}] session={sid[:8]}… | {elapsed}ms | {len(reply)} chars")

        return {
            "reply": reply,
            "session_id": sid,
            "session_title": session.get("title", "New Conversation"),
            "mode": mode,
            "elapsed_ms": elapsed,
            "files_processed": [f["name"] for f in file_contents]
        }

    except EnvironmentError as e:
        log.error(f"Config error: {e}")
        return JSONResponse(status_code=500, content={
            "reply": "⚠️ Server configuration error. GROQ_API_KEY may not be set.",
            "error": str(e)
        })

    except RuntimeError as e:
        log.error(f"Runtime error: {e}")
        return JSONResponse(status_code=502, content={
            "reply": "⚠️ The AI service is currently unavailable. Please try again in a moment.",
            "error": str(e)
        })

    except requests.exceptions.Timeout:
        log.error("Groq request timed out")
        return JSONResponse(status_code=504, content={
            "reply": "⏱️ The request timed out. The model may be busy — please try again.",
        })

    except Exception as e:
        log.exception(f"Unexpected error: {e}")
        return JSONResponse(status_code=500, content={
            "reply": "⚠️ An unexpected server error occurred.",
            "error": str(e)
        })